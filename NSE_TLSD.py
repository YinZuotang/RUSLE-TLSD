# coding=utf-8
# -*- coding:gbk-*-
import arcpy
import os
import sys
import codecs
import pandas as pd
import numpy as np
import openpyxl as op

arcpy.CheckOutExtension("Spatial")
arcpy.env.overwriteOutput = True

ST_A = sys.argv[1]
ST_B = sys.argv[2]
BT = sys.argv[3]
AMAX = sys.argv[4]
DEPMAX = sys.argv[5]
M = sys.argv[6]

with arcpy.EnvManager(scratchWorkspace=r"D:\临时工作空间",
                      workspace=r'D:\AD'):
    StudyArea = r"研究区"
    AD_List = arcpy.ListRasters('*', 'tif')
    StaArea_List = arcpy.ListFeatureClasses("*", "Polygon")
    def nse(shice, moni):
        shice_array = np.array(shice)
        moni_array = np.array(moni)
        Nash_Su = 1 - (np.sum((shice_array - moni_array) ** 2)) / np.sum((shice_array - np.mean(shice_array)) ** 2)
        return Nash_Su

    def getFileName(Path, suffix):
        input_template_All = []
        f_list = os.listdir(Path)  
        for i in f_list:
            if os.path.splitext(i)[1] == suffix:
                input_template_All.append(i)
        return input_template_All

    sczdz_df = pd.read_excel(r"D:\站点.xlsx", header=None)
    sczdz_array = np.array(sczdz_df)
    sczdz_list = sczdz_array.tolist()
    Station = [站点列表]
    Nash_List = [0] * 27
    Nash_List[0] = ST_A
    Nash_List[1] = ST_B
    Nash_List[2] = BT
    Nash_List[3] = AMAX
    Nash_List[4] = DEPMAX
    for AD in AD_List:
        for StaArea in StaArea_List:
            if StaArea[:5]+AD[:6] not in [删选掉一部分没有实测值的数据]:
                StaAreaAD_Name = "D:\\Sta\\" + StaArea[:5] + AD
                StaAreaAD = arcpy.sa.ExtractByMask(AD, StaArea)
                StaAreaAD.save(StaAreaAD_Name)
                print(StaAreaAD_Name)
            else:
                print(StaArea[:5]+AD[:6]+"无站点实测值")

    TifName_List = getFileName("D://Sta", ".tif")
    for TifStaName in [   ]:
        StaTifList = [  ]
        for TifName in TifName_List:
            if TifName[:5] == TifStaName:
                StaTifList.append(TifName)
        TifList=["D://Sta//" + str(i) for i in StaTifList]
        outStatFile = "D://BandTXT//"+TifStaName+".txt"
        arcpy.sa.BandCollectionStats(TifList, outStatFile)

    JZ_MN = []
    JZ_SC = []
    JY_MN = []
    JY_SC = []
    for s in range(   ):
        InFilename = "D://BandTXT//" + Station[s] + ".txt"
        TxtFile = codecs.open(InFilename, mode='r', encoding='utf-8')
        line = TxtFile.readline()
        Txtlist = []
        OneList = []
        RemoveNullList = []
        monistr = []
        moni = []
        while line:
            a = line.split(" ") 
            Txtlist.append(a)
            line = TxtFile.readline()
        Txtlist = Txtlist[6:] 
        Txtlist = Txtlist[:len(Txtlist) - 1] 
        OneList = [j for item in Txtlist for j in item]
        RemoveNullList = [li for li in OneList if li != '']
        mo = 4
        monistr = RemoveNullList[mo - 1::mo + 1]
        moni = list(map(float, monistr))  
        shice_list = sczdz_list[s]
        shice = [SL for SL in shice_list if SL == SL] 
        if s+1 in [   ]:
            JZ_MN.extend(moni)
            JZ_SC.extend(shice)
        elif s+1 in [   ]:
            JY_MN.extend(moni)
            JY_SC.extend(shice)
        Nash = nse(shice, moni)
        Nash_List[s + 5] = Nash
    Nash_List[25] = nse(JZ_SC, JZ_MN)
    Nash_List[26] = nse(JY_SC, JY_MN)

    outputname = op.load_workbook("结果.xlsx")
    sheet = outputname["Sheet1"]
    headername = [参数和站点的精度等表头信息]
    hang = int(M)
    for i in range(0, len(Nash_List)):
        sheet.cell(1, i+1, headername[i])
        sheet.cell(hang, i+1, Nash_List[i])
    outputname.save("结果.xlsx")